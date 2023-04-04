from openpyxl import load_workbook

class Course:
    def __init__(self, course_id, name, description):
        self.course_id = course_id
        self.name = name
        self.description = description

    @staticmethod
    def print_course_trainee_mapping():
        # we cant use the way we done for the method below as that uses dictionary and the problem is that the key would be the courseid but we have multiple re entries of the key courseid
        # and you cant have the same key in a dict as this doesnt make sense so every time we see a key again it would just overwrite th previous value (value is trainee id in this context)
        # of the key to the one it encountered at the moment so when we print the dictionary it would not show us all trainee ids associated with a courseid it woulkd just
        # show the most recent trainee id associated with a courseid
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        course_trainees_mapping_ws = wb["MappingCourseTrainee"]
        current_row = 1
        for row in course_trainees_mapping_ws.iter_rows():
            if course_trainees_mapping_ws.cell(row=current_row,column=1).value == "CourseID":  # skip the first row as we dont want to print that its just headers
                current_row += 1
                continue
            print("course id:", course_trainees_mapping_ws.cell(row=current_row, column=1).value, "| trainee id:",course_trainees_mapping_ws.cell(row=current_row, column=2).value)
            current_row += 1

    @staticmethod
    def get_courses_data():
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        courses_ws = wb["CourseDetails"]

        data = {}

        # https://stackoverflow.com/questions/55319851/how-to-loop-through-each-row-in-excel-spreadsheet-using-openpyxl
        for key, *values in courses_ws.iter_rows(min_row=2):
            data[key.value] = [v.value for v in values]

        return data

    @staticmethod
    def check_course_id(course_id):
        id_list = list(Course.get_courses_data().keys())
        return course_id in id_list

    def add_course(self):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        courses_ws = wb["CourseDetails"]

        courses_ws.append([self.course_id, self.name, self.description])
        wb.save(filename=workbook_name)

    @staticmethod
    def add_course_trainee_mapping(course_id,trainee_id):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        course_trainees_mapping_ws = wb["MappingCourseTrainee"]
        course_trainees_mapping_ws.append([course_id,trainee_id])
        wb.save(filename=workbook_name)

    @staticmethod
    def remove_course(course_id):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        courses_ws = wb["CourseDetails"]

        current_row = 1
        for key, *values in courses_ws.iter_rows():
            # print(f"the key.value is {key.value} but the key is {key}")
            if key.value == course_id:
                courses_ws.delete_rows(current_row, 1)
                wb.save(filename=workbook_name)
                break
            current_row += 1  # at the end of each iteration add 1 to the row number as were moving on to the next row

        # we also need to remove all rows of the with the course id in the course/trainee mappings sheet
        course_trainees_mapping_ws = wb["MappingCourseTrainee"]
        current_row = 1
        for row in course_trainees_mapping_ws.iter_rows():
            if course_trainees_mapping_ws.cell(row=current_row,column=1).value == course_id:
                course_trainees_mapping_ws.delete_rows(current_row, 1)
                wb.save(filename=workbook_name)
                continue # as we don't want to add 1 to current row as we deleted a row so if we on row 4 and we delete row 4 the next row is also row 4
            current_row += 1

        # we also need to remove all rows of the with the course id in the course/trainer mappings sheet
        course_trainer_mapping_ws = wb["MappingCourseTrainer"]
        current_row = 1
        for row in course_trainer_mapping_ws.iter_rows():
            if course_trainer_mapping_ws.cell(row=current_row,column=1).value == course_id:
                course_trainer_mapping_ws.delete_rows(current_row, 1)
                wb.save(filename=workbook_name)
                continue  # as we don't want to add 1 to current row as we deleted a row so if we on row 4 and we delete row 4 the next row is also row 4 as the current tow becomes 3
            current_row += 1

        # lastly we also need to remove all rows of the with the course id in the course/manager mappings sheet
        course_manager_mapping_ws = wb["MappingCourseManager"]
        current_row = 1
        for row in course_manager_mapping_ws.iter_rows():
            if course_manager_mapping_ws.cell(row=current_row,column=1).value == course_id:
                course_manager_mapping_ws.delete_rows(current_row, 1)
                wb.save(filename=workbook_name)
                continue  # as we don't want to add 1 to current row as we deleted a row so if we on row 4 and we delete row 4 the next row is also row 4 as the current tow becomes 3
            current_row += 1

    def update_course(self, original_id):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        courses_ws = wb["CourseDetails"]
        current_row = 1
        for row in courses_ws.iter_rows():
            if courses_ws.cell(row=current_row, column=1).value == original_id:
                courses_ws.cell(row=current_row, column=1).value = self.course_id
                courses_ws.cell(row=current_row, column=2).value = self.name
                courses_ws.cell(row=current_row, column=3).value = self.description
                wb.save(filename=workbook_name)
                break
            current_row += 1

        # we also need to update the course/trainee mapping table if the course id has changed:
        if original_id != self.course_id:
            course_trainees_mapping_ws = wb["MappingCourseTrainee"]
            current_row = 1
            for row in course_trainees_mapping_ws.iter_rows():
                if course_trainees_mapping_ws.cell(row=current_row, column=1).value == original_id:
                    course_trainees_mapping_ws.cell(row=current_row, column=1).value = self.course_id
                    wb.save(filename=workbook_name)
                    # no break as could be multiple
                current_row += 1

        # also we need to update the course/trainer mapping table if the course id has changed:
            course_trainer_mapping_ws = wb["MappingCourseTrainer"]
            current_row = 1
            for row in course_trainer_mapping_ws.iter_rows():
                if course_trainer_mapping_ws.cell(row=current_row,column=1).value == original_id:  # email is on the second column
                    course_trainer_mapping_ws.cell(row=current_row, column=1).value = self.course_id
                    wb.save(filename=workbook_name)
                    # no break as could be multiple
                current_row += 1

        # lastly we need to update the course/manager mapping table if the course id has changed:
            course_manager_mapping_ws = wb["MappingCourseManager"]
            current_row = 1
            for row in course_manager_mapping_ws.iter_rows():
                if course_manager_mapping_ws.cell(row=current_row,column=1).value == original_id:
                    course_manager_mapping_ws.cell(row=current_row, column=1).value = self.course_id
                    wb.save(filename=workbook_name)
                    # no break as could be multiple
                current_row += 1



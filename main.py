from trainees import Trainee
from courses import Course
from trainers import Trainer
from managers import Manager

from datetime import date

from openpyxl import load_workbook


def mark_attendance():
    workbook_name = 'Details.xlsx'
    wb = load_workbook(workbook_name)
    today = str(date.today().strftime("%d_%m_%Y")) # todays date (excel does not allow / in sheet names so we will just use _ instead)

    # get a valid course id
    course_id = input("Enter the id of the course you wish to mark attendance for:\n")
    while not Course.check_course_id(course_id):
        print("This id does not exist!")
        print("Valid course ids:", list(Course.get_courses_data().keys()))
        course_id = input("Enter the id of the course you wish to mark attendance for:\n")

    # get the trainer names associated (if any) with the course
    trainer_names = Trainer.get_course_trainers(course_id)
    valid = False
    if len(trainer_names) > 1:
        print(f"This course has multiple trainers!: {trainer_names}")
        trainer_name = input("Please enter your trainer name to confirm the trainer")
        while trainer_name not in trainer_names:
            print(f"Invalid!\nValid trainers:",trainer_names)
            trainer_name = input("Please enter your trainer name to confirm the trainer")
        valid = True

    elif len(trainer_names) == 1:  # we only have 1 trainer so the name is just the first value in list
        trainer_name = trainer_names[0]
        valid = True

    else: # else it means the course has no trainers:
        print("Error this course has no trainers, please add a trainer to the course to mark attendance!")
        return

    if valid:  # if we have a valid course with a trainer associated with it:
        # create the actual sheet in excel
        sheet_name = today + "_" + course_id
        wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

        # get start and end times:
        print(f"Add session for course {course_id} and date {today}:")
        start_time = input("Enter session start time:\n")
        end_time = input("Enter session end time:\n")
        title = "Course: " + course_id + " | Date: " + today + " | Session Time: " + start_time + " to " + end_time + " | Trainer: " + trainer_name
        ws.merge_cells('A1:C1')
        ws.cell(row=1, column=1).value = title # we cant do append as for some reason that adds onto the next line not on the merged line

        # changing width of A1
        length_of_column = len(ws['A1'].value)
        ws.column_dimensions['A'].width = length_of_column

        ws.append(["Name","TraineeID","Status"]) # headers
        ids, names = Trainee.get_course_trainees(course_id)
        if len(ids) > 0:
            for i in range(len(ids)):  # so for each index
                print(f"\nFor the following student: {ids[i]} - {names[i]}")
                status = None
                while status not in ['P', 'A']:
                    status = input("Enter P for present or A for absent\n")
                ws.append([names[i], ids[i], status])
            wb.save(filename=workbook_name)  # dont forget to save!
            print("Attendance submitted")
        else:
            print("Error no students enrolled in this course!")
            return


if __name__ == "__main__":
    choice = None
    while choice != 24:
        print("Enter a valid number corresponding to your desired action:")
        print("1 - Add a trainee")
        print("2 - Remove a trainee")
        print("3 - Update a trainees details")
        print("4 - View all trainee details")
        print("5 - Add a course")
        print("6 - Remove a course")
        print("7 - Update a courses details")
        print("8 - View all courses details")
        print("9 - Add a course/trainee mapping")
        print("10 - View course and trainee mappings")
        print("11 - Add a trainer")
        print("12 - Remove a trainer")
        print("13 - Update a trainers details")
        print("14 - View all trainers details")
        print("15 - Add a course/trainer mapping")
        print("16 - View course and trainer mappings")
        print("17 - Add a manager")
        print("18 - Remove a manager")
        print("19 - Update a managers details")
        print("20 - View all managers details")
        print("21 - Add a course/manager mapping")
        print("22 - View all course and manager mappings")
        print("23 - Submit attendance report")
        print("24 - Quit")

        choice = int(input())

        if choice == 1:
            trainee_id = int(input("Enter trainee id (must be unique):\n"))
            while Trainee.check_trainee_id(trainee_id):
                print("This id is already taken!")
                trainee_id = int(input("Enter trainee id (must be unique):\n"))
            name = input("Enter trainee name:\n")
            degree = input("Enter the trainees degree:\n")
            work_experience = input("Enter any relevant work experience the trainee has:\n")

            trainee = Trainee(trainee_id, name, degree, work_experience)
            trainee.add_trainee()
            print("Trainee added successfully")

        elif choice == 2:
            trainee_id = int(input("Enter the id of the trainee you wish to remove:\n"))
            while not Trainee.check_trainee_id(trainee_id):
                print("This id does not exist!")
                print("Valid trainee ids:", list(Trainee.get_trainees_data().keys()))
                trainee_id = int(input("Enter the id of the trainee you wish to remove:\n"))
            Trainee.remove_trainee(trainee_id)
            print(f"Trainee with id {trainee_id} successfully removed")

        elif choice == 3:
            original_trainee_id = int(input("Enter the trainee id of the trainee you wish to edit:\n"))
            while not Trainee.check_trainee_id(original_trainee_id):
                print("This id does not exist!")
                print("Valid trainee ids:", list(Trainee.get_trainees_data().keys()))
                original_trainee_id = int(input("Enter the trainee id of the trainee you wish to edit:\n"))

            print("Enter the same values if you wish to not change a field:")
            new_trainee_id = int(input("Enter trainee id (must be unique):\n"))
            while Trainee.check_trainee_id(new_trainee_id) and new_trainee_id != original_trainee_id: # if the user enters a id thats is diffrent from the og id and is already taken
                print("This id is already taken!")
                new_trainee_id = int(input("Enter trainee id (must be unique):\n"))
            name = input("Enter trainee name:\n")
            degree = input("Enter the trainees degree:\n")
            work_experience = input("Enter any relevant work experience the trainee has:\n")

            trainee = Trainee(new_trainee_id, name, degree, work_experience)
            trainee.update_trainee(original_trainee_id)
            print("Trainee details successfully updated")

        elif choice == 4:
            trainees = Trainee.get_trainees_data()
            for key, value in trainees.items():
                print(f"Details of trainee with trainee id: {key}")
                print(f"  Name: {value[0]}\n  Degree: {value[1]}\n  Work Experience: {value[2]}\n")

        elif choice == 5:
            course_id = input("Enter course id (must be unique):\n")
            while Course.check_course_id(course_id):
                print("This id is already taken!")
                course_id = input("Enter course id (must be unique):\n")

            name = input("Enter the course name:\n")
            description = input("Enter course description:\n")

            course = Course(course_id,name,description)
            course.add_course()
            print("Course added successfully")

        elif choice == 6:
            course_id = input("Enter the course id of the course you wish to remove:\n")
            while not Course.check_course_id(course_id):
                print("This id does not exist!")
                print("Valid course ids:",list(Course.get_courses_data().keys()))
                course_id = input("Enter the course id of the course you wish to remove:\n")
            Course.remove_course(course_id)
            print(f"Course with id {course_id} successfully removed")

        elif choice == 7:
            original_course_id = input("Enter the course id of the course you wish to edit:\n")
            while not Course.check_course_id(original_course_id):
                print("This id does not exist!")
                print("Valid course ids:",list(Course.get_courses_data().keys()))
                original_course_id = input("Enter the course id of the course you wish to edit:\n")

            print("Enter the same values if you wish to not change a field:")
            new_course_id = input("Enter course id (must be unique):\n")
            while Course.check_course_id(new_course_id) and new_course_id != original_course_id:  # if the user enters a id thats is diffrent from the og id and is already taken
                print("This id is already taken!")
                new_course_id = input("Enter course id (must be unique):\n")
            name = input("Enter course name:\n")
            description = input("Enter the courses description:\n")

            course = Course(new_course_id, name, description)
            course.update_course(original_course_id)
            print("Course details successfully updated")

        elif choice == 8:
            courses = Course.get_courses_data()
            for key, value in courses.items():
                print(f"Details of course with course id: {key}")
                print(f"  Name: {value[0]}\n  Description: {value[1]}\n")

        elif choice == 9:
            course_id = input("Enter the course id:\n")
            while not Course.check_course_id(course_id): # if the course id doesnt exist (cant enroll into course that's not there)
                print("Not a valid course id!")
                print("Valid course ids:",list(Course.get_courses_data().keys()))
                course_id = input("Enter course id:\n")

            trainee_id = int(input("Enter trainee id:\n"))
            while not Trainee.check_trainee_id(trainee_id):
                print("This id doesnt exist!")
                print("Valid trainee ids:", list(Trainee.get_trainees_data().keys()))
                trainee_id = int(input("Enter trainee id:\n"))

            Course.add_course_trainee_mapping(course_id,trainee_id)
            print("Mapping successfully added")

        elif choice == 10:
            Course.print_course_trainee_mapping()
            print()

        elif choice == 11:
            trainer_email = input("Enter the trainer email (must be unique)\n")
            while Trainer.check_trainer_email(trainer_email):
                print("This email is already taken!")
                trainer_email = input("Enter the trainer email (must be unique)\n")
            name = input("Enter trainer name:\n")
            phone = input("Enter the trainers phone number:\n")

            trainer = Trainer(trainer_email, name, phone)
            trainer.add_trainer()
            print("Trainer added successfully")

        elif choice == 12:
            trainer_email = input("Enter the trainer email of the trainer you wish to remove:\n")
            while not Trainer.check_trainer_email(trainer_email):
                print("This email does not exist!")
                print("Valid trainer emails:", list(Trainer.get_trainers_data().keys()))
                trainer_email = input("Enter the trainer email of the trainer you wish to remove:\n")
            Trainer.remove_trainer(trainer_email)
            print(f"Trainer with email {trainer_email} successfully removed")

        elif choice == 13:
            original_trainer_email = input("Enter the trainer email of the trainer you wish to edit:\n")
            while not Trainer.check_trainer_email(original_trainer_email):
                print("This email does not exist!")
                print("Valid trainer emails:", list(Trainer.get_trainers_data().keys()))
                original_trainer_email = input("Enter the trainer email of the trainer you wish to edit:\n")

            print("Enter the same values if you wish to not change a field:")
            new_trainer_email = input("Enter trainer email (must be unique):\n")
            while Trainer.check_trainer_email(new_trainer_email) and new_trainer_email != original_trainer_email:  # if the user enters a email thats is diffrent from the og email and is already taken
                print("This email is already taken!")
                new_trainer_email = input("Enter trainer email (must be unique):\n")
            name = input("Enter trainer name:\n")
            phone = input("Enter trainers phone number:\n")
            trainer = Trainer(new_trainer_email, name, phone)
            trainer.update_trainer(original_trainer_email)
            print("Trainer details successfully updated")

        elif choice == 14:
            trainers = Trainer.get_trainers_data()
            for key, value in trainers.items():
                print(f"Details of trainer with email id: {key}")
                print(f"  Name: {value[0]}\n  Phone Number: {value[1]}\n")

        elif choice == 15:
            course_id = input("Enter the course id:\n")
            while not Course.check_course_id(course_id):  # if the course id doesnt exist (cant enroll into course that's not there)
                print("Not a valid course id!")
                print("Valid course ids:", list(Course.get_courses_data().keys()))
                course_id = input("Enter course id:\n")

            email = input("Enter email of trainer:\n")
            while not Trainer.check_trainer_email(email):
                print("This email doesnt exist!")
                print("Valid emails:", list(Trainer.get_trainers_data().keys()))
                email = input("Enter email of trainer:\n")

            Trainer.add_course_trainer_mapping(course_id, email)
            print("Mapping successfully added")

        elif choice == 16:
            Trainer.print_course_trainers_mapping()
            print()

        elif choice == 17:
            manager_email = input("Enter the manager email (must be unique)\n")
            while Manager.check_manager_email(manager_email):
                print("This email is already taken!")
                manager_email = input("Enter the manager email (must be unique)\n")
            name = input("Enter manager name:\n")
            phone = input("Enter the managers phone number:\n")

            manager = Manager(manager_email, name, phone)
            manager.add_manager()
            print("Manager added successfully")

        elif choice == 18:
            manager_email = input("Enter the email of the manager you wish to remove:\n")
            while not Manager.check_manager_email(manager_email):
                print("This email does not exist!")
                print("Valid manager emails:", list(Manager.get_managers_data().keys()))
                manager_email = input("Enter the email of the manager you wish to remove:\n")
            Manager.remove_manager(manager_email)
            print(f"Manager with email {manager_email} successfully removed")

        elif choice == 19:
            original_manager_email = input("Enter the email of the manager you wish to edit:\n")
            while not Manager.check_manager_email(original_manager_email):
                print("This email does not exist!")
                print("Valid manager emails:", list(Manager.get_managers_data().keys()))
                original_manager_email = input("Enter the email of the manager you wish to edit:\n")

            print("Enter the same values if you wish to not change a field:")
            new_manager_email = input("Enter manager email (must be unique):\n")
            while Manager.check_manager_email(new_manager_email) and new_manager_email != original_manager_email:  # if the user enters a email thats is diffrent from the og email and is already taken
                print("This email is already taken!")
                new_manager_email = input("Enter manager email (must be unique):\n")
            name = input("Enter manager name:\n")
            phone = input("Enter managers phone number:\n")
            manager = Manager(new_manager_email, name, phone)
            manager.update_manager(original_manager_email)
            print("Manager details successfully updated")

        elif choice == 20:
            managers = Manager.get_managers_data()
            for key, value in managers.items():
                print(f"Details of manager with email id: {key}")
                print(f"  Name: {value[0]}\n  Phone Number: {value[1]}\n")

        elif choice == 21:
            course_id = input("Enter the course id:\n")
            while not Course.check_course_id(
                    course_id):  # if the course id doesnt exist (cant enroll into course that's not there)
                print("Not a valid course id!")
                print("Valid course ids:", list(Course.get_courses_data().keys()))
                course_id = input("Enter course id:\n")

            email = input("Enter email of manager:\n")
            while not Manager.check_manager_email(email):
                print("This email doesnt exist!")
                print("Valid emails:", list(Manager.get_managers_data().keys()))
                email = input("Enter email of manager:\n")

            Manager.add_course_manager_mapping(course_id, email)
            print("Mapping successfully added")

        elif choice == 22:
            Manager.print_course_managers_mapping()
            print()

        elif choice == 23:
            mark_attendance()
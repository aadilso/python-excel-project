from openpyxl import load_workbook


class Manager:
    def __init__(self, email, name, phone):
        self.email = email
        self.name = name
        self.phone = phone

    @staticmethod
    def print_course_managers_mapping():
        # we cant use the way we done for the method below as that uses dictionary and the problem is that the key would be the courseid but we have multiple re entries of the key courseid
        # and you cant have the same key in a dict as this doesnt make sense so every time we see a key again it would just overwrite th previous value (value is managerid in this context)
        # of the key to the one it encountered at the moment so when we print the dictionary it would not show us all manager ids associated with a course id it woulkd just
        # show the most recent manager id associated with a courseid

        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        course_manager_mapping_ws = wb["MappingCourseManager"]
        current_row = 1
        for row in course_manager_mapping_ws.iter_rows():
            if course_manager_mapping_ws.cell(row=current_row,column=1).value == "CourseID":  # skip the first row as we dont want to print that its just headers
                current_row += 1
                continue
            print("course id:", course_manager_mapping_ws.cell(row=current_row, column=1).value, "| manager email id:",course_manager_mapping_ws.cell(row=current_row, column=2).value)
            current_row += 1

    @staticmethod
    def get_managers_data():
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        managers_ws = wb["ListOfManagers"]

        data = {}

        # https://stackoverflow.com/questions/55319851/how-to-loop-through-each-row-in-excel-spreadsheet-using-openpyxl
        for key, *values in managers_ws.iter_rows(min_row=2): # min row = 2 as we want to skip the headings row
            data[key.value] = [v.value for v in values]

        return data

    @staticmethod
    def check_manager_email(manager_email):
        email_list = list(Manager.get_managers_data().keys())
        return manager_email in email_list

    def add_manager(self):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        managers_ws = wb["ListOfManagers"]

        # note when modifying a excel file the file must be closed when we run the program
        managers_ws.append([self.email, self.name, self.phone])
        wb.save(filename=workbook_name)

    @staticmethod
    def add_course_manager_mapping(course_id, email_id):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        course_manager_mapping_ws = wb["MappingCourseManager"]
        course_manager_mapping_ws.append([course_id, email_id])
        wb.save(filename=workbook_name)

    @staticmethod
    def remove_manager(email_id):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        managers_ws = wb["ListOfManagers"]

        current_row = 1
        for key, *values in managers_ws.iter_rows():
            # print(f"the key.value is {key.value} but the key is {key}")
            if key.value == email_id:
                managers_ws.delete_rows(current_row, 1)
                wb.save(filename=workbook_name)
                break
            current_row += 1  # at the end of each iteration add 1 to the row number as were moving on to the next row

        # we also need to remove all rows of the with the course id in the course/manager mappings sheet
        course_manager_mapping_ws = wb["MappingCourseManager"]
        current_row = 1
        for row in course_manager_mapping_ws.iter_rows():
            if course_manager_mapping_ws.cell(row=current_row,column=2).value == email_id:  # in the mappings table the second column is the email field
                course_manager_mapping_ws.delete_rows(current_row, 1)
                wb.save(filename=workbook_name)
                continue  # as we don't want to add 1 to current row as we deleted a row so if we on row 4 and we delete row 4 the next row is also row 4 as the current tow becomes 3
            current_row += 1

    def update_manager(self, original_email):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        managers_ws = wb["ListOfManagers"]
        current_row = 1
        for row in managers_ws.iter_rows():
            if managers_ws.cell(row=current_row, column=1).value == original_email:
                managers_ws.cell(row=current_row, column=1).value = self.email
                managers_ws.cell(row=current_row, column=2).value = self.name
                managers_ws.cell(row=current_row, column=3).value = self.phone
                wb.save(filename=workbook_name)
                break
            current_row += 1

        # we also need to update the course/manager mapping table if the email has changed:
        if original_email != self.email:
            course_manager_mapping_ws = wb["MappingCourseManager"]
            current_row = 1
            for row in course_manager_mapping_ws.iter_rows():
                if course_manager_mapping_ws.cell(row=current_row, column=2).value == original_email:  # email is on the second column
                    course_manager_mapping_ws.cell(row=current_row, column=2).value = self.email
                    wb.save(filename=workbook_name)
                    # no break as could be multiple
                current_row += 1

from openpyxl import Workbook, load_workbook


class Trainee:
    def __init__(self, trainee_id, name, degree, work_exp):
        self.trainee_id = trainee_id
        self.name = name
        self.degree = degree
        self.work_exp = work_exp

    @staticmethod
    def get_trainees_data():
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        trainees_ws = wb["ListOfTrainees"]

        data = {}

        # https://stackoverflow.com/questions/55319851/how-to-loop-through-each-row-in-excel-spreadsheet-using-openpyxl
        for key, *values in trainees_ws.iter_rows(min_row=2):
            data[key.value] = [v.value for v in values]

        return data

    @staticmethod
    # returns true if id exists, false if not
    def check_trainee_id(trainee_id):
        id_list = list(Trainee.get_trainees_data().keys())
        return trainee_id in id_list

    @staticmethod
    # method that returns a list of trainee ids associated with a course and a list of trainee names associated with a course
    def get_course_trainees(course_id):
        course_trainee_ids = []
        course_trainee_names = []

        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        course_trainee_mapping_ws = wb["MappingCourseTrainee"]
        current_row = 1
        for row in course_trainee_mapping_ws.iter_rows():
            if course_trainee_mapping_ws.cell(row=current_row,column=1).value == course_id:  # if the courseid matches, add the trainee id to our lst
                course_trainee_ids.append(course_trainee_mapping_ws.cell(row=current_row, column=2).value)
            current_row += 1

        all_trainees = Trainee.get_trainees_data()
        for key, value in all_trainees.items():
            if key in course_trainee_ids:
                course_trainee_names.append(value[0])

        # in python we can return multiple values (in this case lists) like so:
        return course_trainee_ids, course_trainee_names

    def add_trainee(self):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        trainees_ws = wb["ListOfTrainees"]

        # note when modifying a excel file the file must be closed when we run the program, we also ave to save if we forget te canes wont reflect in te file
        trainees_ws.append([self.trainee_id,self.name,self.degree,self.work_exp])
        wb.save(filename=workbook_name)

    @staticmethod
    def remove_trainee(trainee_id):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        trainees_ws = wb["ListOfTrainees"]
        current_row = 1
        for key, *values in trainees_ws.iter_rows():
            # print(f"the key.value is {key.value} but the key is {key}")
            if key.value == trainee_id:
                trainees_ws.delete_rows(current_row, 1)
                wb.save(filename=workbook_name)
                break
            current_row += 1  # at the end of each iteration add 1 to the row number as were moving on to the next row

        # we also need to remove any entries of the trainee from our mappings sheet
        course_trainees_mapping_ws = wb["MappingCourseTrainee"]
        current_row = 1
        for row in course_trainees_mapping_ws.iter_rows():
            if course_trainees_mapping_ws.cell(row=current_row, column=2).value == trainee_id:
                course_trainees_mapping_ws.delete_rows(current_row, 1)
                wb.save(filename=workbook_name)
                continue # as we don't want to add 1 to current row as we deleted a row so if we on row 4 and we delete row 4 the next row is also row 4
            current_row += 1

    def update_trainee(self,original_id):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        trainees_ws = wb["ListOfTrainees"]
        current_row = 1
        for row in trainees_ws.iter_rows():
            if trainees_ws.cell(row=current_row, column=1).value == original_id:
                trainees_ws.cell(row=current_row, column=1).value = self.trainee_id
                trainees_ws.cell(row=current_row, column=2).value = self.name
                trainees_ws.cell(row=current_row, column=3).value = self.degree
                trainees_ws.cell(row=current_row, column=4).value = self.work_exp
                wb.save(filename=workbook_name)
                break
            current_row += 1

        # also need to update the course/trainee mappings sheet if the trainee id has been changed:
        if original_id != self.trainee_id:
            course_trainees_mapping_ws = wb["MappingCourseTrainee"]
            current_row = 1
            for row in course_trainees_mapping_ws.iter_rows():
                if course_trainees_mapping_ws.cell(row=current_row, column=2).value == original_id:
                    course_trainees_mapping_ws.cell(row=current_row, column=2).value = self.trainee_id
                    wb.save(filename=workbook_name)
                    # don't break as there could be multiple instances of that trainee id in the sheet!
                current_row += 1

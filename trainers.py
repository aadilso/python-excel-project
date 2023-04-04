from openpyxl import load_workbook


class Trainer:
    def __init__(self, email, name, phone):
        self.email = email
        self.name = name
        self.phone = phone

    @staticmethod
    def print_course_trainers_mapping():
        # we cant use the way we done for the method below as that uses dictionary and the problem is that the key would be the courseid but we have multiple re entries of the key courseid
        # and you cant have the same key in a dict as this doesnt make sense so every time we see a key again it would just overwrite th previous value (value is trainee id in this context)
        # of the key to the one it encountered at the moment so when we print the dictionary it would not show us all trainee ids associated with a courseid it woulkd just
        # show the most recent trainee id associated with a courseid
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        course_trainer_mapping_ws = wb["MappingCourseTrainer"]
        current_row = 1
        for row in course_trainer_mapping_ws.iter_rows():
            if course_trainer_mapping_ws.cell(row=current_row,column=1).value == "CourseID":  # skip the first row as we dont want to print that its just headers
                current_row += 1
                continue
            print("course id:", course_trainer_mapping_ws.cell(row=current_row, column=1).value, "| trainer email id:",course_trainer_mapping_ws.cell(row=current_row, column=2).value)
            current_row += 1

    @staticmethod
    def get_trainers_data():
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        courses_ws = wb["ListOfTrainers"]

        data = {}

        # https://stackoverflow.com/questions/55319851/how-to-loop-through-each-row-in-excel-spreadsheet-using-openpyxl
        for key, *values in courses_ws.iter_rows(min_row=2):
            data[key.value] = [v.value for v in values]

        return data

    @staticmethod
    def check_trainer_email(trainer_email):
        email_list = list(Trainer.get_trainers_data().keys())
        return trainer_email in email_list

    @staticmethod
    # method returns a list of the trainer name/s for a specific course id
    def get_course_trainers(course_id):
        course_trainer_emails = []
        course_trainer_names = []

        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        course_trainer_mapping_ws = wb["MappingCourseTrainer"]
        current_row = 1
        for row in course_trainer_mapping_ws.iter_rows():
            if course_trainer_mapping_ws.cell(row=current_row,column=1).value == course_id:  # if the courseid matches add the trainer email to our lst
                course_trainer_emails.append(course_trainer_mapping_ws.cell(row=current_row, column=2).value)
            current_row += 1

        all_trainers = Trainer.get_trainers_data()
        for key, value in all_trainers.items():
            if key in course_trainer_emails:
                course_trainer_names.append(value[0])

        return course_trainer_names

    def add_trainer(self):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        trainer_ws = wb["ListOfTrainers"]

        # note when modifying a excel file the file must be closed when we run the program
        trainer_ws.append([self.email, self.name, self.phone])
        wb.save(filename=workbook_name)

    @staticmethod
    def add_course_trainer_mapping(course_id,email_id):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        course_trainers_mapping_ws = wb["MappingCourseTrainer"]
        course_trainers_mapping_ws.append([course_id, email_id])
        wb.save(filename=workbook_name)

    @staticmethod
    def remove_trainer(email_id):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        trainer_ws = wb["ListOfTrainers"]

        current_row = 1
        for key, *values in trainer_ws.iter_rows():
            # print(f"the key.value is {key.value} but the key is {key}")
            if key.value == email_id:
                trainer_ws.delete_rows(current_row, 1)
                wb.save(filename=workbook_name)
                break
            current_row += 1  # at the end of each iteration add 1 to the row number as were moving on to the next row

            # we also need to remove all rows of the with the course id in the course/trainer mappings sheet
        course_trainer_mapping_ws = wb["MappingCourseTrainer"]
        current_row = 1
        for row in course_trainer_mapping_ws.iter_rows():
            if course_trainer_mapping_ws.cell(row=current_row, column=2).value == email_id: # in the mappings table the second column is the email field
                course_trainer_mapping_ws.delete_rows(current_row, 1)
                wb.save(filename=workbook_name)
                continue  # as we don't want to add 1 to current row as we deleted a row so if we on row 4 and we delete row 4 the next row is also row 4 as the current tow becomes 3
            current_row += 1

    def update_trainer(self, original_email):
        workbook_name = 'Details.xlsx'
        wb = load_workbook(workbook_name)
        trainer_ws = wb["ListOfTrainers"]
        current_row = 1
        for row in trainer_ws.iter_rows():
            if trainer_ws.cell(row=current_row, column=1).value == original_email:
                trainer_ws.cell(row=current_row, column=1).value = self.email
                trainer_ws.cell(row=current_row, column=2).value = self.name
                trainer_ws.cell(row=current_row, column=3).value = self.phone
                wb.save(filename=workbook_name)
                break
            current_row += 1

        # we also need to update the course/trainer mapping table if the email has changed:
        if original_email != self.email:
            course_trainer_mapping_ws = wb["MappingCourseTrainer"]
            current_row = 1
            for row in course_trainer_mapping_ws.iter_rows():
                if course_trainer_mapping_ws.cell(row=current_row, column=2).value == original_email: # email is on the second column
                    course_trainer_mapping_ws.cell(row=current_row, column=2).value = self.email
                    wb.save(filename=workbook_name)
                    # no break as could be multiple
                current_row += 1
